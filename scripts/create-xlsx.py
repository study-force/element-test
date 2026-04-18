from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="1E293B")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_font = Font(name="Arial", size=10)
cell_align = Alignment(vertical="center", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
axis_fill = PatternFill("solid", fgColor="FEF3C7")
mixed_fill = PatternFill("solid", fgColor="F0F9FF")

# === Sheet 1 ===
ws1 = wb.active
ws1.title = "문항 목록"

for col, h in enumerate(["프레임ID", "프레임 테마", "블록 유형", "차원", "#", "키", "극성", "문항 텍스트"], 1):
    c = ws1.cell(row=1, column=col, value=h)
    c.font, c.fill, c.alignment, c.border = header_font, header_fill, header_align, thin_border

frames = [
    (7, "원리 vs 경험", "순수축 θ1", "", [
        ("경험_1", "경험", "실제 사례나 경험이 있어야 내용이 머릿속에 들어온다"),
        ("이론_2", "이론", "왜 그런지 이유까지 알아야 받아들여진다"),
        ("경험_2", "경험", "원리를 알지 못해도 수행에 어려움이 없으면 상관없다"),
        ("이론_1", "이론", "어떤 현상을 보면 왜 그렇게 되는 건지 원리가 궁금하다"),
    ]),
    (8, "행동 vs 사고", "순수축 θ2", "", [
        ("사고_2", "사고", "결정을 내리기 전에 여러 가능성을 충분히 따져봐야 마음이 편하다"),
        ("실행_1", "실행", "직접 해보거나 말로 설명하면 생각이 더 잘 정리되는 편이다"),
        ("사고_1", "사고", "행동하기 전에 머릿속에서 충분히 정리가 되어야 시작할 수 있다"),
        ("실행_2", "실행", "몸을 움직이거나 뭔가를 하면서 배울 때 집중이 더 잘 된다"),
    ]),
    (1, "뭔가를 이해할 때 나는", "혼합", "내면·감정", [
        ("경험사고", "", "개념이 내 경험이나 감정과 연결돼야 이해된 느낌이 든다"),
        ("이론실행", "", "개념이 이해되면 문제를 풀어 바로 확인해보고 싶어진다"),
        ("이론사고", "", "개념이 충분히 이해된 후 문제까지 모두 풀어야 안심이 된다"),
        ("경험실행", "", "개념 이해가 완벽하지 않아도 일단 문제를 풀면서 익히는 게 편하다"),
    ]),
    (2, "처음 접하는 것들 앞에서", "혼합", "외면·행동", [
        ("이론사고", "", "새로운 게임이나 앱을 접하면, 이용하기 전 매뉴얼을 보고 전체 기능을 세세히 파악한다"),
        ("경험실행", "", "새로운 게임이나 앱을 접하면, 일단 이것저것 눌러보며 익힌다"),
        ("경험사고", "", "새로운 게임이나 앱을 접하면, 체험판을 해보거나 후기들을 살펴본 후 이용한다"),
        ("이론실행", "", "새로운 게임이나 앱을 접하면, 주요 방법만 파악하고 이용한다"),
    ]),
    (3, "가장 몰입되는 순간", "혼합", "내면·동기", [
        ("경험실행", "", "새로운 것을 배우거나 변화가 있을 때 집중력이 강해진다"),
        ("이론사고", "", "하나를 오래 깊이 파고들수록 점점 더 집중된다"),
        ("이론실행", "", "원리가 이해되는 순간 쾌감이 생기면서 집중력이 높아진다"),
        ("경험사고", "", "혼자 천천히 되짚어볼 때 오히려 더 많은 것이 보인다"),
    ]),
    (4, "벽에 부딪혔을 때", "혼합", "내면·전략", [
        ("경험사고", "", "막히면 비슷한 문제나 예시를 찾아보면서 실마리를 얻는 편이다"),
        ("이론사고", "", "막힌 부분이 해결되지 않으면 다음으로 넘어가기가 불편하다"),
        ("경험실행", "", "막히면 일단 다른 문제부터 풀고 나중에 다시 돌아온다"),
        ("이론실행", "", "막힌 부분은 원리부터 다시 확인하면 대부분 해결된다"),
    ]),
    (5, "모둠 활동을 할 때 나는", "혼합", "외면·타인(긍정)", [
        ("이론사고", "", "모둠 활동을 할 때, 내가 잘 해낼 수 있는 역할을 신중하게 선택한다"),
        ("이론실행", "", "모둠 활동을 할 때, 전체 방향을 제시하고 이끄는 편이다"),
        ("경험사고", "", "모둠 활동을 할 때, 조원들에게 어울리는 역할이 무엇인지 생각한다"),
        ("경험실행", "", "모둠 활동을 할 때, 내가 하고 싶은 역할을 빠르게 정하는 편이다"),
    ]),
    (6, "공부할 때 나는", "혼합", "외면·타인(부정)", [
        ("경험실행", "", "공부할 때, 계획을 세우기보다는 생각 나는대로 한다"),
        ("경험사고", "", "공부할 때, 이것저것 신경쓰느라 한 가지 공부를 깊이 하기 어렵다"),
        ("이론실행", "", "공부할 때, 큰 줄기만 이해되면 넘어간다"),
        ("이론사고", "", "공부할 때, 꼼꼼하게 정리하느라 시간이 오래 걸린다"),
    ]),
]

row = 2
for fid, theme, btype, dim, stmts in frames:
    fill = axis_fill if "순수축" in btype else mixed_fill
    for i, (key, pole, text) in enumerate(stmts):
        vals = [f"F{fid}", theme, btype, dim, i + 1, key, pole if pole else "-", text]
        aligns = [center_align, cell_align, center_align, center_align, center_align, center_align, center_align, cell_align]
        for col, (v, a) in enumerate(zip(vals, aligns), 1):
            c = ws1.cell(row=row, column=col, value=v)
            c.font, c.alignment, c.border, c.fill = cell_font, a, thin_border, fill
        row += 1

ws1.column_dimensions["A"].width = 10
ws1.column_dimensions["B"].width = 24
ws1.column_dimensions["C"].width = 12
ws1.column_dimensions["D"].width = 16
ws1.column_dimensions["E"].width = 5
ws1.column_dimensions["F"].width = 10
ws1.column_dimensions["G"].width = 7
ws1.column_dimensions["H"].width = 60
ws1.auto_filter.ref = f"A1:H{row - 1}"
ws1.freeze_panes = "A2"

# === Sheet 2 ===
ws2 = wb.create_sheet("유형 정보")

for col, h in enumerate(["유형 키", "이모지", "캐릭터명", "유형명", "슬로건", "부제", "강점", "약점"], 1):
    c = ws2.cell(row=1, column=col, value=h)
    c.font, c.fill, c.alignment, c.border = header_font, header_fill, header_align, thin_border

types_data = [
    ("이론실행", "\U0001f525", "불의 정복자", "이론-실행형", "원리를 꿰뚫어 바로 움직이는 학습자", "원리이해 + 추진력", "논리력·분석력, 추진력, 문제해결력", "단순 암기, 반복 학습, 세부 디테일", "EDF5FA"),
    ("경험실행", "\U0001f32a\ufe0f", "바람의 전령", "경험-실행형", "일단 해보며 길을 찾는 학습자", "호기심 + 실행력", "추진력·도전정신, 높은 적응력, 실행력", "원리 이해 부족, 지속력·완성도", "FFF2ED"),
    ("이론사고", "\U0001faa8", "땅의 수호자", "이론-사고형", "깊이 생각하고 체계를 완성하는 학습자", "계획력 + 지구력", "체계력·계획력, 완성도·지구력, 깊은 이해력", "처리 속도, 유연성·결단력", "FAF5ED"),
    ("경험사고", "\U0001f4a7", "물의 정령", "경험-사고형", "느끼고 관찰하며 의미를 찾는 학습자", "상상력 + 섬세함", "암기력·꼼꼼함, 공감력·다각적 관점, 풍부한 아이디어", "속도·결단력, 논리 분석", "EDF8FB"),
]

for i, (*vals, bg) in enumerate(types_data):
    fill = PatternFill("solid", fgColor=bg)
    for col, v in enumerate(vals, 1):
        c = ws2.cell(row=i + 2, column=col, value=v)
        c.font, c.alignment, c.border, c.fill = cell_font, cell_align, thin_border, fill

ws2.column_dimensions["A"].width = 10
ws2.column_dimensions["B"].width = 6
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 14
ws2.column_dimensions["E"].width = 32
ws2.column_dimensions["F"].width = 18
ws2.column_dimensions["G"].width = 36
ws2.column_dimensions["H"].width = 30
ws2.freeze_panes = "A2"

out = "C:/Users/이재훈/Desktop/CLAUDE/element-test/ELEMENT_TEST_ITEMS.xlsx"
wb.save(out)
print(f"Saved: {out}")
